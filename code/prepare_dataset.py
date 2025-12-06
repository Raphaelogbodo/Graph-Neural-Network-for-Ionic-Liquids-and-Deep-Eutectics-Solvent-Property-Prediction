''' Installed Packages
#!pip install ilthermopy
#!pip install seaborn
#!pip install rdkit
#!pip install openbabel-wheel
#!pip install pybel
#!conda install -c conda-forge openbabel -y
#!pip install rdkit-pypi
#!pip install pubchempy
#!pip install mordred
#!pip install py3Dmol
#!pip install pyilt2
#!pip install torch torchvision torchaudio --index-url https://download.pytorch.org/whl/cpu
#!pip install torch-scatter torch-sparse torch-cluster torch-spline-conv torch-geometric -f https://data.pyg.org/whl/torch-2.6.0+cpu.html
#!pip install torchmetrics
#!pip install permetrics
#!pip install plotly
#!pip install openpyxl
#!pip install --upgrade kaleido
#!pip install optuna

was not able to install this on windows, try if you want
#!pip install torch-scatter torch-sparse torch-cluster torch-spline-conv
'''
#########################################################################################################################

'''import plotting tools'''
import matplotlib as mpl 
import matplotlib.pyplot as plt
import seaborn as sns 
import requests, os
import pandas as pd
import re
import numpy as np
import pubchempy as pcp
from rdkit import Chem
from rdkit.Chem import rdmolops
from mordred import Calculator, descriptors
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error
import matplotlib.pyplot as plt
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from torch.nn import Linear
from torch_geometric.data import Data, DataLoader
from torch_geometric.nn import GCNConv, global_mean_pool
import ilthermopy as ilt
import py3Dmol
from openbabel import openbabel
from openbabel import pybel
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

pd.plotting.register_matplotlib_converters()
#%matplotlib inline 
pd.set_option('display.max_rows',None)
pd.set_option('display.max_columns',None)
pd.set_option('display.max_colwidth', None)
print(plt.style.available)
mpl.style.use(['seaborn-v0_8-poster'])

mpl.rcParams['xtick.labelsize'] = 14
mpl.rcParams['ytick.labelsize'] = 14
mpl.rcParams['font.weight'] = 'normal' # 'bold'
mpl.rcParams['axes.labelweight'] = 'normal' # 'bold'
mpl.rcParams['axes.labelcolor'] = 'black'
mpl.rcParams['xtick.color'] = 'black'
mpl.rcParams['ytick.color'] = 'black'

print(torch.__version__)
print(torch.version.cuda)

'''Paths for saving figures and data files'''
def make_dir(base, folder):
    path = os.path.join(base, folder)
    os.makedirs(path, exist_ok=True)
    return path

#### Neccessary Paths Definitions 
home_dir = '/Users/rnogbodo/ML/raph_GNN' # replace with your own main folder path
DATA = make_dir(home_dir,'data')
Figures = make_dir(home_dir,'Figures')


'''Dowloding all the ILs Entries from the ILThermo (v2.0) Database'''
df_all = ilt.GetAllEntries()
print(f"Length of IL Thermo Database Entries Retrieved --------> {len(df_all)}")

'''Extract all Entries with phases=Liquid'''
df_all_liquid = df_all[(df_all['phases'] == 'Liquid')]
print(f"Length of IL Thermo Database Entries with phases = Liquid --------> {len(df_all_liquid)}")

'''All Liquid Phase Entry IDs'''
entry_ids = df_all_liquid['id'].values.tolist()

'''Unique number of components'''
unique_num_components = list(df_all_liquid['num_components'].unique())

'''Let create a separate DataFrame for each of the components'''
df_all_liquid_cmp1 = df_all_liquid[(df_all_liquid['num_components'] == 1)]
print(f"Shape of IL Thermo Database Entries with phase = Liquid and num_components = 1 -------> {df_all_liquid_cmp1.shape}\n")

df_all_liquid_cmp2 = df_all_liquid[(df_all_liquid['num_components'] == 2)]
print(f"Shape of IL Thermo Database Entries with phase = Liquid and num_components = 2 -------> {df_all_liquid_cmp2.shape}\n")

df_all_liquid_cmp3 = df_all_liquid[(df_all_liquid['num_components'] == 3)]
print(f"Shape of IL Thermo Database Entries with phase = Liquid and num_components = 3 -------> {df_all_liquid_cmp3.shape}\n")

'''Let us check if they all have entries for all the properties'''
assert df_all_liquid_cmp1['property'].unique().all() == df_all_liquid_cmp1['property'].unique().all() == df_all_liquid_cmp1['property'].unique().all(), 'different properties' 

'''Since they contain same properties, Let us make a list of properties'''
properties = list(df_all_liquid_cmp1['property'].unique())

'''
Let us try identifying unique IL entries Based on either cation or anion cores
These functions --> filter_by_cation_core() and filter_by_anion_core() ----> are used to find the string of interest in each column
'''

def filter_by_cation_core(s, core):
    """
    Checks if the given cation core is in the correct position in a cmp1 string.
    Rules:
    - If split into 2 parts: only check first part
    - If split into more than 2 parts: check first or second part
    """
    if pd.isna(s):
        return False
    parts = s.split()
    parts_lower = [p.lower() for p in parts]
    target = core.lower()
    
    if len(parts) == 2:
        return target in parts_lower[0]
    elif len(parts) > 2:
        return any(target in parts_lower[i] for i in range(2))  # first or second
    else:
        return target in parts_lower[0]

def filter_by_anion_core(s, core):
    """
    Checks if the given cation core is in the correct position in a cmp1 string.
    Rules:
    - If split into 2 parts: only check second part
    - If split into more than 2 parts: check second or more other part
    """
    if pd.isna(s):
        return False
    parts = s.split()
    parts_lower = [p.lower() for p in parts]
    target = core.lower()
    
    if len(parts) == 2:
        return target in parts_lower[1]
    elif len(parts) > 2:
        return any(target in parts_lower[i] for i in range(2,len(parts)))  # secod or third or fourth
    else:
        return False

filter_cores = {'cation':filter_by_cation_core, 'anion':filter_by_anion_core} # This dictionary now contains the instance of the functions above


'''
These functions below get the combined dataframe and a dataframe containing ion cores and their counts
Function process_df_components() only works for a single col variable
Function process_df_components_v2() accepts col variable as a list of columns to scanned from
Function process_df_components_v3() does similar thing as _v2()
'''

def process_df_components(df,named_cores_dict,named_cores_map, col,motif='cation'):
    named_cores = list(named_cores_dict.values())
    named_core_dfs = {core: df[df[col].apply(lambda x: filter_cores[motif](x, cor))] for cor, core in named_cores_dict.items()}
    def get_total():
        tot = 0
        for core in named_cores:
            print(f"{core} ----------------> {named_core_dfs[core].shape}")
            tot += named_core_dfs[core].shape[0]
        print(f"\nRecovered entries ---> {tot}\nMaximum entries in the given data frame ---->{df.shape[0]}\n")
    #get_total()

    '''Concating all the named_cores data frame'''
    df_out = pd.DataFrame(columns=df.columns)
    core_frequency = {motif:[],'count':[]}
    for core in named_cores:
        row, col = named_core_dfs[core].shape
        core_frequency[motif].append(core)
        core_frequency['count'].append(row)
        df_out = pd.concat([df_out, named_core_dfs[core]], axis=0, ignore_index=False)
    df_core_freq = pd.DataFrame(core_frequency)
    
    df_out = df_out.drop_duplicates(subset="id", keep="first").reset_index(drop=True)
    df_out = pd.concat([df_out, df[~df['id'].isin(df_out['id'])]], axis=0)
    print(f"Total entries after concating all named_cores data frames -------> {df_out.shape[0]}\n")

    '''Merging named_cores that are same but named in different ways for the sake of entry counts'''
    df_core_freq[motif] = df_core_freq[motif].replace(named_cores_map) # replace names of old with new
    df_core_freq = df_core_freq.groupby(motif, as_index=False)["count"].sum() # group by motif and sum 
    df_core_freq = df_core_freq[df_core_freq["count"].astype(float) != 0.0]
    df_core_freq = df_core_freq.sort_values("count", ascending=False).reset_index(drop=True) # sort and reset index

    return df_out, df_core_freq

def process_df_components_v2(df,named_cores_dict,named_cores_map, cols,motif='cation'):
    named_cores = list(named_cores_dict.values())
    core_freq = {motif:[],'count':[]}
    df_freq = pd.DataFrame(core_freq)
    df_all_out = pd.DataFrame(columns=df.columns)
    for col in cols:
        named_core_dfs = {core: df[df[col].apply(lambda x: filter_cores[motif](x, cor))] for cor, core in named_cores_dict.items()}
       
        '''Concating all the named_cores data frame'''
        df_out = pd.DataFrame(columns=df.columns)
        core_frequency = {motif:[],'count':[]}
        for core in named_cores:
            row, col = named_core_dfs[core].shape
            core_frequency[motif].append(core)
            core_frequency['count'].append(row)
            df_out = pd.concat([df_out, named_core_dfs[core]], axis=0, ignore_index=False)
        df_core_freq = pd.DataFrame(core_frequency)
        
        df_out = df_out.drop_duplicates(subset="id", keep="first").reset_index(drop=True)
        #df_out = pd.concat([df_out, df[~df['id'].isin(df_out['id'])]], axis=0)
            
        '''Merging named_cores that are same but named in different ways for the sake of entry counts'''
        df_core_freq[motif] = df_core_freq[motif].replace(named_cores_map) # replace names of old with new
        df_core_freq = df_core_freq.groupby(motif, as_index=False)["count"].sum() # group by motif and sum 
        df_core_freq = df_core_freq.sort_values("count", ascending=False).reset_index(drop=True) # sort and reset index

        df_all_out = pd.concat([df_all_out, df_out], axis=0)
        df_freq = pd.concat([df_freq, df_core_freq], axis=0)

    df_all_out = df_all_out.drop_duplicates(subset="id", keep="first").reset_index(drop=True)
    print(f"\nRecovered entries ---> {df_all_out.shape[0]}\nMaximum entries in the given data frame ---->{df.shape[0]}\n")
    df_all_out = pd.concat([df_all_out, df[~df['id'].isin(df_all_out['id'])]], axis=0)
    print(f"Total entries after concating all named_cores data frames -------> {df_all_out.shape[0]}\n")

    df_freq = df_freq.groupby(motif, as_index=False)["count"].sum() # group by motif and sum 
    df_freq = df_freq[df_freq["count"].astype(float) != 0.0]
    df_freq = df_freq.sort_values("count", ascending=False).reset_index(drop=True) # sort and reset index
    
    return df_all_out, df_freq

def process_df_components_v3(df, named_cores_dict, named_cores_map, cols, motif='cation'):
    """
    Process DataFrame to extract and count entries matching named cores across multiple columns.

    Parameters:
    - df: pd.DataFrame containing molecular components.
    - named_cores_dict: dict mapping motif keys to canonical core names.
    - named_cores_map: dict to merge different motif names into canonical names.
    - cols: list of column names to scan for motifs.
    - motif: str label for the type of motif (e.g., 'anion' or 'cation').

    Returns:
    - df_all_out: DataFrame containing all matched rows plus unmatched rows at the end.
    - df_freq: DataFrame with motif names and total counts.
    """
    
    named_cores = list(named_cores_dict.values())
    df_all_out = pd.DataFrame(columns=df.columns)
    freq_records = []

    for col in cols:
        # Filter rows for each core
        core_dfs = {core: df[df[col].apply(lambda x: filter_cores[motif](x, cor))] 
                    for cor, core in named_cores_dict.items()}

        # Concatenate all core DataFrames for this column
        df_out = pd.concat(core_dfs.values(), axis=0)
        
        # Count rows per core
        for core in named_cores:
            row_count = core_dfs[core].shape[0]
            if row_count > 0:
                freq_records.append({motif: core, 'count': row_count})

        # Append unique rows to final output
        df_all_out = pd.concat([df_all_out, df_out], axis=0)

    # Remove duplicates based on 'id'
    df_all_out = df_all_out.drop_duplicates(subset='id', keep='first').reset_index(drop=True)

    # Add unmatched rows
    df_all_out = pd.concat([df_all_out, df[~df['id'].isin(df_all_out['id'])]], axis=0).reset_index(drop=True)

    # Build frequency DataFrame
    df_freq = pd.DataFrame(freq_records)
    if not df_freq.empty:
        # Apply the mapping and sum counts
        df_freq[motif] = df_freq[motif].replace(named_cores_map)
        df_freq = df_freq.groupby(motif, as_index=False)['count'].sum()
        df_freq = df_freq[df_freq['count'].astype(float) != 0.0]
        df_freq = df_freq.sort_values('count', ascending=False).reset_index(drop=True)

    print(f"\nRecovered entries ---> {df_all_out.shape[0]}\nMaximum entries in the given DataFrame ---> {df.shape[0]}\n")
    print(f"Total entries after concatenating all named core DataFrames -------> {df_all_out.shape[0]}\n")

    return df_all_out, df_freq

'''
Ion cores Dictionary and their possible maps
Notice that ammonium and aminium are same core, choline and cholinium are same core, we need to combine them
'''

cation_cores_dict ={
    'imidazol':'imidazolium',         # imidazolium ring, e.g., 1-butyl-3-methylimidazolium
    'pyridin':'pyridinium',           # pyridinium ring, e.g., 1-hexylpyridinium
    'pyrrolidin':'pyrrolidinium',     # pyrrolidinium, e.g., N-butyl-N-methylpyrrolidinium
    'pyrrolidone':'pyrrolidone',      # pyrrolidone, eg., N-methylpyrrolidone tetrafluoroborate
    'piperidin':'piperidinium',           # piperidinium, e.g., N-butyl-N-methylpiperidinium
    'pyrazol':'pyrazolium',        # pyrazolium eg., 1-ethyl-2-methylpyrazolium dicyanamide
    'phosphonium':'phosphonium',      # tetraalkylphosphonium
    'ammonium':'ammonium',            # tetraalkylammonium
    'guanid':'guanidinium',      # 1,1,3,3-tetramethylguanidinium
    'morpholin':'morpholinium',    # N-butyl-N-methylmorpholinium
    'sulfonium':'sulfonium',          # trialkylsulfonium
    'thiazol':'thiazolium',           # thiazolium ring
    'triazol':'triazolium',           # triazolium ring
    'quinolinium':'quinolinium',      # quinolinium ring
    'isoquinolinium':'isoquinolinium',# isoquinolinium ring
    'aminium':'aminium',              # aminium
    'thiophen':'thiophenium',      # thiophenium
    'choline':'choline',              # choline
    'cholinium':'cholinium',          # cholinium
    'amine':'amine',                  # amine
    'amid':'amidium',                 # amidium  eg., N,N-dimethylacetamide trifluoroacetate
    'azepanium':'azepanium',          # azepanium eg.,1-(4-sulfobutyl)azepanium hydrogen sulfate
    'propenium':'propenium',          # propenium eg., tris(dibutylamino)cyclopropenium dicyanamide
    'tropinium':'tropinium',          # tropinium eg., N-butyltropinium glycinate
    'valinate': 'valinate',           # valinate eg., isobutyl L-valinate dodecyl sulfate
    'piperazinium':'piperazinium',    # piperazinium eg., N-ethylpiperazinium propionate
    'aniline':'anilinium',            # anilinium eg., aniline formate
    'alani':'alanine',                # alanine eg., L-alanine, 1-methylethyl ester, dodecyl sulfate
    'imizodalium':'imizodalium',      # imizodalium eg., 1-allyl-3-methylimizodalium acetate
    'azepinium':'azepinium',          # azepinium  eg., 1-decyl-2,3,4,6,7,8,9,10-octahydropyrimido[1,2-a]azepinium thiocyanate
    'pydridinium':'pydridinium',      # pydridinium  eg., 1-octyl-3-methylpydridinium bis(trifluoromethylsulfonyl)imide
    'thianium':'thianium',            # thianium   eg., 1-butylthianium bis(trifluoromethylsulfonyl)imide
    'thiouronium':'thiouronium',      # thiouronium  eg., S-octyl-1,1,3,3-tetramethylthiouronium bis(trifluoromethylsulfonyl)imide
    'leucine':'leucine',              # leucine  eg., L-(+)-N-butylleucine ethyl ester tetrafluoroborate
    'quinuclidinium':'quinuclidinium',# quinuclidinium  eg., 1-hexylquinuclidinium bis((trifluoromethyl)sulfonyl)amide
    'pyrimid':'pyrimidinium',       # pyrimidinium  eg., 1-methyl-1,2,3,4,5,6,7,8-octahydropyrimido[1,2-a]pyrimidin-5-ium acetate
    'non-5-enium':'non-5-enium',      # non-5-enium  eg., 1,5-diazabicyclo[4.3.0]non-5-enium 3-aminopyrazolide
    'nonanium':'nonanium',
    'lithium': 'lithium',
    'aminoethanol':'aminoethanol',
    'potassium':'potassium'
    
}

anion_cores_dict = {
    'amide':'amide',
    'imide':'imide',
    'anide':'anide',
    'borate':'borate',
    'nitrate':'nitrate',
    'sulfate':'sulfate',
    'sulfonate':'sulfonate',
    'chloride':'chloride',
    'bromide':'bromide',
    'phosphate':'phosphate',
    'acetate':'acetate',
    'iodide':'iodide',
    'propionate':'propionate',
    'butyrate':'butyrate',
    'oate':'alkanoate',
    'formate':'formate',
    'levulinate':'levulinate',
    'borohydride':'borohydride',
    'saccharinate':'saccharinate',
    'tricyanamenthane':'tricyanamethane',
    'thiocyanate':'thiocyanate',
    'alaninate':'amino_acid',
    'imid':'imid',
    'lactate':'lactate',
    'phosphinate':'phosphinate',
    'prolinate':'prolinate',
    'valinate':'valinate',
    'glycinate':'glycinate',
    'triazolide':'triazolide',
    'pyrazolide':'pyrazolide',
    'imidazolide':'imidazolide',
    'tetrazolide':'tetrazolide',
    'enoate':'enoate',
    'oleate':'oleate',
    'serinate':'serinate',
    'taurinate':'taurinate',
    '-ide':'ide',
    'tosylate':'tosylate',
    'carbonate':'carbonate',
    'phenolate':'phenolate',
    'pyrrolide':'pyrrolide',
    'indazolide':'indazolide',
    '-yloxyl':'yloxyl',
    'tryptophanate':'tryptophanate',
    'aspartate':'aspartate',
    'glycine':'glycine',
    'chromate':'chromate',
    'threoninate':'threoninate',
    'cysteinate':'cysteinate',
    'lysinate':'lysinate',
    'taurate':'taurate',
    'glutamate':'glutamate',
    'salicylate':'salicylate',
    'succinate':'succinate',
    'carboxylate':'carboxylate',
     'glycolate':'glycolate',
    'stearate':'stearate',
    'phosphonate':'phosphonate',
    'acrylate':'acrylate',
    'cinnamate':'cinnamate',
}


# mapping of equivalents
cation_map = {
    "aminium": "ammonium",
    "choline": "ammonium",
    'cholinium':'ammonium',
    'amine':'ammonium',
    'imizodalium':'imidazolium',
    'pydridinium':'pyridinium',
    'aminoethanol':'ammonium',
    'isoquinolinium':'quinolinium',
}

anion_map = {
    'imide':'amide',
    'propionate':'alkanoate',
    'borohydride':'borate',
    'imid':'amide',
    'anide':'amide',
    'prolinate':'amino_acid',
    'valinate':'amino_acid',
    'glycinate':'amino_acid',
    'serinate':'amino_acid',
    'aspartate':'amino_acid',
    'glycine':'amino_acid',
    'threoninate':'amino_acid',
    'cysteinate':'amino_acid',
    'lysinate':'amino_acid',
    'glutamate':'amino_acid',
    'tryptophanate':'amino_acid',
    'levulinate':'carboxylate',
    'formate':'carboxylate',
    'acetate':'carboxylate',
    'butyrate':'carboxylate',
    'lactate':'carboxylate',
    'taurinate':'carboxylate',
    'salicylate':'carboxylate',
    'glycolate':'carboxylate',
    'stearate':'carboxylate',
    'acrylate':'carboxylate',
    'cinnamate':'carboxylate',
    'succinate':'carboxylate',
    'taurate':'carboxylate',
}

'''
You can check elements of cmp1 not in cmp2
'''
cmp1 = set(df_all_liquid['cmp1'].unique())
cmp2 = set(df_all_liquid['cmp2'].unique())

# disjoint set (symmetric difference)
disjoint = cmp2 - cmp1

'''Filtering techniques'''
#df_all_liquid_cmp1[df_all_liquid_cmp1['cmp1'].str.contains('L-valinate', case=False, na=False)
#df_all_liquid_cmp1[df_all_liquid_cmp1['cmp1'].apply(lambda x: filter_cores['cation'](x, 'phospho'))].shape
#df_all_liquid_cmp1[df_all_liquid_cmp1['cmp1'].apply(lambda x: filter_by_anion_core(x, 'bromide'))].head()


'''Some graphical statistics of the database'''
# Extract 4-digit years from the 'reference' column
df_new = df_all.copy()
df_new['year'] = df_new['reference'].str.extract(r'\((\d{4})\)')
df_years = df_new.dropna(subset=['year']).copy()
df_years['year'] = df_years['year'].astype(int)

df_year_counts = df_years['year'].value_counts().sort_index().reset_index()
df_year_counts.columns = ['year', 'count']

# Example: df_year_counts has ['year', 'count']
df_year_counts['year'] = df_year_counts['year'].astype(int)
df_year_counts = df_year_counts.sort_values('year').reset_index(drop=True)

# Map years to categorical positions
positions = np.arange(len(df_year_counts))
df_year_counts['pos'] = positions

fig, ax = plt.subplots(figsize=(16, 10))

# Plot bars
sns.barplot(x='pos', y='count', data=df_year_counts, ax=ax, color='royalblue', edgecolor='black')

# Smart tick selection: show roughly ~15 ticks
max_ticks = 26
step = max(1, len(df_year_counts) // max_ticks)
tick_positions = df_year_counts['pos'][::step]
tick_labels = df_year_counts['year'][::step]

ax.set_xticks(tick_positions)
ax.set_xticklabels(tick_labels, rotation=70, ha='right', fontsize=20, color='black')

# Labels and styling
ax.set_ylabel('Count (log scale)', fontsize=30, fontweight='bold', color='black')
ax.set_xlabel('Year', fontsize=30, fontweight='bold', color='black')
ax.set_title('Studies on ILs', fontsize=30, fontweight='bold', pad=10, color='black')
ax.tick_params(axis='both', labelsize=20, colors='black')
ax.set_yscale('log')

plt.tight_layout()
plt.savefig(os.path.join(Figures, 'pub-per-year.svg'),
            dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()


motif_cat = 'cation'
cols=['cmp1','cmp2','cmp3']
df_cation_cores, df_cation_core_freq = process_df_components_v3(df_all_liquid,cation_cores_dict,cation_map,cols,motif=motif_cat)

df_cation_core_freq[motif_cat] = df_cation_core_freq[motif_cat].apply(lambda x: x.capitalize())
sns.barplot(y=motif_cat,x='count', data=df_cation_core_freq[df_cation_core_freq['count'] > 0], edgecolor='black', color='blue')
#plt.xticks(rotation=90)   # rotate x-axis labels
#plt.title(f'{motif_cat.capitalize()} Cores Distributions for Pure,\nBinary, and Ternary mixtures',fontsize=30, fontweight='bold',pad=10)

plt.ylabel('Cation', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'cation-core-all.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()


motif_an = 'anion'
cols=['cmp1','cmp2','cmp3']
df_anion_cores, df_anion_core_freq = process_df_components_v3(df_all_liquid,anion_cores_dict,anion_map,cols,motif=motif_an)

df_anion_core_freq[motif_an] = df_anion_core_freq[motif_an].apply(lambda x: x.capitalize())
sns.barplot(y=motif_an,x='count', data=df_anion_core_freq[df_anion_core_freq['count'] > 0], edgecolor='black',color='blue')
#plt.xticks(rotation=90)   # rotate x-axis labels
#plt.title(f'{motif_an.capitalize()} Cores Distributions for Pure,\nBinary, and Ternary mixtures',fontsize=30, fontweight='bold',pad=10)
plt.ylabel('Anion', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'anion-core-all.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

'''Distribution of Cation Cores for different number of components, 1, 2, or 3'''
motif_cat = 'cation'
cols=['cmp1','cmp2','cmp3']
df_cation_cores, df_cation_core_freq = process_df_components_v3(df_all_liquid_cmp1,cation_cores_dict,cation_map,cols,motif=motif_cat)

df_cation_core_freq[motif_cat] = df_cation_core_freq[motif_cat].apply(lambda x: x.capitalize())
sns.barplot(y=motif_cat,x='count', data=df_cation_core_freq[df_cation_core_freq['count'] > 0 ], edgecolor='black', color='blue')
#plt.xticks(rotation=90)   # rotate x-axis labels
plt.title(f'{motif_cat.capitalize()} Cores Distributions for Pure ILs', fontsize=30, fontweight='bold', pad=10)

plt.ylabel('Cation', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'cation-core-pure.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

motif_cat = 'cation'
cols=['cmp1','cmp2','cmp3']
df_cation_cores, df_cation_core_freq = process_df_components_v3(df_all_liquid_cmp2,cation_cores_dict,cation_map,cols,motif=motif_cat)

df_cation_core_freq[motif_cat] = df_cation_core_freq[motif_cat].apply(lambda x: x.capitalize())
sns.barplot(y=motif_cat,x='count', data=df_cation_core_freq[df_cation_core_freq['count'] > 0 ], edgecolor='black', color='blue')
#plt.xticks(rotation=90)   # rotate x-axis labels
plt.title(f'{motif_cat.capitalize()} Cores Distributions for Binary Mixtures',fontsize=30, fontweight='bold',pad=10 )

plt.ylabel('Cation', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'cation-core-binary.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

motif_cat = 'cation'
cols=['cmp1','cmp2','cmp3']
df_cation_cores, df_cation_core_freq = process_df_components_v3(df_all_liquid_cmp3,cation_cores_dict,cation_map,cols,motif=motif_cat)

df_cation_core_freq[motif_cat] = df_cation_core_freq[motif_cat].apply(lambda x: x.capitalize())
sns.barplot(y=motif_cat,x='count', data=df_cation_core_freq[df_cation_core_freq['count'] > 0 ], edgecolor='black', color='blue')
#plt.xticks(rotation=90)   # rotate x-axis labels
plt.title(f'{motif_cat.capitalize()} Cores Distributions for Ternary Mixtures', fontsize=30, fontweight='bold', pad=10)

plt.ylabel('Cation', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'cation-core-ternary.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

''''Distribution of Anion Cores for different number of components, 1, 2, or 3'''
motif_an = 'anion'
cols=['cmp1','cmp2','cmp3']
df_anion_cores, df_anion_core_freq = process_df_components_v3(df_all_liquid_cmp1,anion_cores_dict,anion_map,cols,motif=motif_an)

df_anion_core_freq[motif_an] = df_anion_core_freq[motif_an].apply(lambda x: x.capitalize())
sns.barplot(y=motif_an,x='count', data=df_anion_core_freq[df_anion_core_freq['count'] > 0], edgecolor='black', color='blue')
#plt.xticks(rotation=90)   # rotate x-axis labels
plt.title(f'{motif_an.capitalize()} Cores Distributions for Pure IL',fontsize=30, fontweight='bold', pad=10)

plt.ylabel('Anion', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'anion-core-pure.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

motif_an = 'anion'
cols=['cmp1','cmp2','cmp3']
df_anion_cores, df_anion_core_freq = process_df_components_v3(df_all_liquid_cmp2,anion_cores_dict,anion_map,cols,motif=motif_an)

df_anion_core_freq[motif_an] = df_anion_core_freq[motif_an].apply(lambda x: x.capitalize())
sns.barplot(y=motif_an,x='count', data=df_anion_core_freq[df_anion_core_freq['count'] > 0], edgecolor='black', color='blue')
#plt.xticks(rotation=90)   # rotate x-axis labels
plt.title(f'{motif_an.capitalize()} Cores Distributions for Binary Mixtures', fontsize=30, fontweight='bold', pad=10)
plt.ylabel('Anion', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'anion-core-binary.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

motif_an = 'anion'
cols=['cmp1','cmp2','cmp3']
df_anion_cores, df_anion_core_freq = process_df_components_v3(df_all_liquid_cmp3,anion_cores_dict,anion_map,cols,motif=motif_an)

df_anion_core_freq[motif_an] = df_anion_core_freq[motif_an].apply(lambda x: x.capitalize())
sns.barplot(y=motif_an,x='count', data=df_anion_core_freq[df_anion_core_freq['count'] > 0], edgecolor='black', color='blue')
plt.title(f'{motif_an.capitalize()} Cores Distributions for Ternary Mixtures', fontsize=30, fontweight='bold',pad=10)
plt.ylabel('Anion', fontsize=30, fontweight='bold')
plt.xlabel('Count (log scale)', fontsize=30, fontweight='bold')
plt.xscale('log')
plt.savefig(os.path.join(Figures, 'anion-core-ternary.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

'''
By now, we already have idea of the distribution of cation cores and number of components
We have duly selected IL that are in the liquid phase and separated them according to number of components
one component ---> df_all_liquid_cmp1
two component ---> df_all_liquid_cmp2
three component ---> df_all_liquid_cmp3

Let us start looking into specific properties of interest whether in pure, binary or ternary mixtures
'''
items = {'property':[], 'all_liquid':[], 'liquid_cmp1':[], 'liquid_cmp2':[], 'liquid_cmp3':[]}
for prop in properties:
    num_all_L = df_all_liquid[df_all_liquid['property'] == prop].shape[0]
    num_all_L_cmp1 = df_all_liquid_cmp1[df_all_liquid_cmp1['property'] == prop].shape[0]
    num_all_L_cmp2 = df_all_liquid_cmp2[df_all_liquid_cmp2['property'] == prop].shape[0]
    num_all_L_cmp3 = df_all_liquid_cmp3[df_all_liquid_cmp3['property'] == prop].shape[0]
    #print(f"{prop} -----> all comps:{num_all_L},  comps 1:{num_all_L_cmp1}")
    items['property'].append(prop) ; items['all_liquid'].append(num_all_L) ; items['liquid_cmp1'].append(num_all_L_cmp1)
    items['liquid_cmp2'].append(num_all_L_cmp2) ; items['liquid_cmp3'].append(num_all_L_cmp3)
df_prop_counts = pd.DataFrame(items)

df_prop_counts.index = df_prop_counts['property']
df_prop_counts.index.name = None
df_prop_counts.drop(columns=['property'], inplace=True)

rename_index = {
    "Heat capacity at constant pressure": "Heat Capacity P",
    "Heat capacity at constant volume": "Heat Capacity V",
    "Viscosity": "Viscosity",
    "Density": "Density",
    "Speed of sound": "Sound Speed",
    "Isothermal compressibility": "Compressibility T",
    "Electrical conductivity": "Electrical Cond.",
    "Isobaric coefficient of volume expansion": "Volume Exp. Coeff.",
    "Refractive index": "Refractive Index",
    "Surface tension liquid-gas": "Surface Tension",
    "Thermal conductivity": "Thermal Cond.",
    "Enthalpy function {H(T)-H(0)}/T": "Enthalpy Function",
    "Entropy": "Entropy",
    "Enthalpy": "Enthalpy",
    "Self diffusion coefficient": "Self Diffusion",
    "Relative permittivity": "Relative Permittivity",
    "Thermal diffusivity": "Thermal Diff."
}

df_prop_counts_renamed = df_prop_counts.rename(index=rename_index)

'''Plotting publication counts against properties for pure, binary, and ternary mixtures'''
colors = ['black', 'red', 'blue']
xticks = np.arange(len(df_prop_counts_renamed.index.tolist()))
df_prop_counts_renamed[['liquid_cmp1', 'liquid_cmp2','liquid_cmp3']].plot(kind='line', marker='o', figsize=(12,6), linestyle='None', color=colors)
plt.xticks(ticks=np.arange(len(df_prop_counts_renamed)), labels=df_prop_counts_renamed.index, rotation=70)
plt.ylabel('Publications', fontsize=30, fontweight='bold')
plt.xlabel('Property', fontsize=30, fontweight='bold')
plt.legend(labels=['Pure ILs', 'Binary Mixtures', 'Ternary Mixtures'],fontsize=20)
plt.yscale('log')
plt.savefig(os.path.join(Figures, 'property-pub.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

'''Let me quickly get information for other phases combined'''
df_non_liquid = df_all[df_all['phases'] != 'Liquid']
property_shortnames = {
    'Heat capacity at vapor saturation pressure': 'Heat capacity sat P',
    'Heat capacity at constant pressure': 'Heat capacity P',
    'Electrical conductivity': 'Conductivity',
    'Equilibrium pressure': 'Equil. pressure',
    'Density': 'Density',
    'Enthalpy': 'Enthalpy',
    'Entropy': 'Entropy',
    'Enthalpy function {H(T)-H(0)}/T': 'Enthalpy/T',
    'Viscosity': 'Viscosity',
    'Speed of sound': 'Sound speed',
    'Surface tension liquid-gas': 'Surface tension',
    'Refractive index': 'Refractive index',
    'Equilibrium temperature': 'Equil. temperature',
    'Self diffusion coefficient': 'Self diffusion',
    'Thermal conductivity': 'Thermal conductivity',
    'Enthalpy of vaporization or sublimation': 'Enthalpy vapor',
    'Enthalpy of transition or fusion': 'Enthalpy fusion',
    'Normal melting temperature': 'Melting temp',
    'Normal boiling temperature': 'Boiling temp',
    'Vapor or sublimation pressure': 'Vapor pressure',
    'Composition at phase equilibrium': 'Equil. composition',
    'Interfacial tension': 'Interfacial tension',
    'Osmotic coefficient': 'Osmotic coefficient',
    'Activity': 'Activity',
    'Binary diffusion coefficient': 'Binary diffusion',
    'Tracer diffusion coefficient': 'Tracer diffusion',
    'Thermal diffusivity': 'Thermal diffusivity',
    'Excess enthalpy': 'Excess enthalpy',
    'Apparent enthalpy': 'Apparent enthalpy',
    "Henry's Law constant": 'Henry constant',
    'Critical pressure': 'Critical pressure',
    'Excess volume': 'Excess volume',
    'Upper consolute pressure': 'Upper consolute P',
    'Eutectic composition': 'Eutectic composition',
    'Eutectic temperature': 'Eutectic temp',
    'Monotectic temperature': 'Monotectic temp',
    'Enthalpy of solution': 'Enthalpy solution',
    'Upper consolute temperature': 'Upper consolute T',
    'Upper consolute composition': 'Upper consolute comp',
    'Critical temperature': 'Critical temperature',
    'Tieline': 'Tieline',
    'Ostwald coefficient': 'Ostwald coefficient',
    'Partial molar enthalpy ': 'Partial enthalpy',
    'Lower consolute temperature': 'Lower consolute T'
}

items = {'property':[], 'all_liquid':[]}
for prop in property_shortnames.keys():
    num_all_non_L = df_non_liquid[df_non_liquid['property'] == prop].shape[0]
    items['property'].append(prop) ; items['all_liquid'].append(num_all_non_L)
df_counts = pd.DataFrame(items)

df_counts.index = df_counts['property']
df_counts.index.name = None
df_counts.drop(columns=['property'], inplace=True)

df_counts_renamed = df_counts.rename(index=property_shortnames)
colors = ['black']
xticks = np.arange(len(df_prop_counts_renamed.index.tolist()))
df_counts_renamed.plot(kind='line', marker='o',markeredgecolor='black',markerfacecolor='royalblue', figsize=(12,6), color=colors)
plt.xticks(ticks=np.arange(len(df_counts_renamed)), labels=df_counts_renamed.index, rotation=90)
plt.ylabel('Publications', fontsize=30, fontweight='bold')
plt.xlabel('Property', fontsize=30, fontweight='bold')
plt.legend(labels=['Other Phases'],fontsize=20)
plt.yscale('log')
plt.savefig(os.path.join(Figures, 'property-pub-other-phases.svg'),dpi=600, bbox_inches='tight',facecolor='white')
#plt.show()

'''Now we need to get into the actual data of the properties of interest'''

properties_of_interest = {
    'Heat capacity at constant pressure': 'Heat_capacity_Cp',
    'Viscosity':'Viscosity',
    'Density':'Density',
    'Electrical conductivity':'Conductivity',
    'Refractive index':'Refractive_index',
    'Surface tension liquid-gas':'Surface_tension',
}

entries_id_cmp1_visc = df_all_liquid_cmp1[df_all_liquid_cmp1['property'] =='Viscosity']['id'].tolist()
entries_id_cmp1_cp = df_all_liquid_cmp1[df_all_liquid_cmp1['property'] =='Heat capacity at constant pressure']['id'].tolist()
entries_id_cmp1_dens = df_all_liquid_cmp1[df_all_liquid_cmp1['property'] =='Density']['id'].tolist()
entries_id_cmp1_cond = df_all_liquid_cmp1[df_all_liquid_cmp1['property'] =='Electrical conductivity']['id'].tolist()
entries_id_cmp1_refrac = df_all_liquid_cmp1[df_all_liquid_cmp1['property'] =='Refractive index']['id'].tolist()
entries_id_cmp1_surftens = df_all_liquid_cmp1[df_all_liquid_cmp1['property'] =='Surface tension liquid-gas']['id'].tolist()


def format_reference_robust(ref: str) -> str:
    """
    Format a reference string into 'FirstAuthor et al. (Year) Journal'.
    Handles single/multiple authors and variations in reference formatting.
    Parameters
    ----------
    ref : str
        Full reference string.
    Returns
    -------
    str
        Formatted citation, e.g., 'Fan et al. (2021) J. Mol. Liq.'
    """
    # 1. Extract first author
    first_author_match = re.match(r'\s*([A-Za-z\-]+)', ref)
    first_author = first_author_match.group(1) if first_author_match else "Unknown"
    # 2. Determine if multiple authors exist
    multiple_authors = ";" in ref or "," in ref.split(")")[0]  # check before year
    author_str = f"{first_author} et al." if multiple_authors else first_author
    # 3. Extract year (4-digit number in parentheses or after comma)
    year_match = re.search(r"\((\d{4})\)|, (\d{4})", ref)
    if year_match:
        year = year_match.group(1) if year_match.group(1) else year_match.group(2)
    else:
        year = "Unknown"
    # 4. Extract journal (text after year up to first volume/page number)
    # Accept letters, spaces, periods, & some punctuation
    journal_match = re.search(r"\)\s*([A-Za-z0-9\.\s\-\&]+?)(?=\s+\d|,|$)", ref)
    journal = journal_match.group(1).strip() if journal_match else "Unknown Journal"

    return f"{author_str} ({year}) {journal}"

def separate_cat_from_an_similes(smile):
    def get_charges(smile):
        mol = Chem.MolFromSmiles(smile)
        mol_mol = Chem.MolToMolBlock(mol)
        bel_mol = pybel.readstring('mol', mol_mol)
        ob_charge_model = openbabel.OBChargeModel.FindType('gasteiger')
        ob_charge_model.ComputeCharges(bel_mol.OBMol)
        charges = [bel_mol.OBMol.GetAtom(i+1).GetFormalCharge() for i, atom in enumerate(mol.GetAtoms())]
        return charges
    smiles = smile.split('.')
    cat_smiles = []  ; an_smiles = []
    if len(smiles) >= 2:
        for smile in smiles:
            charges = get_charges(smile)
            if sum(charges) > 0:
                cat_smiles.append(smile)
            elif sum(charges) < 0:
                an_smiles.append(smile)
        cat_smile = '.'.join(cat_smiles)
        an_smile = '.'.join(an_smiles)
        return (cat_smile, an_smile)
    elif len(smiles) < 2:
        return smile

def get_entry_data_colnames(patterns, entry):
    matched_keys = {}
    for key, val in entry.header.items():
        for name, pat in patterns.items():
            if pat.search(val):
                matched_keys[name] = key
    return matched_keys
	
def get_database_data(columns, must_get, patterns, entries_id_cmp1):
    data = {key:[] for key in columns}  
    for idx in entries_id_cmp1:
        entry = ilt.GetEntry(idx)
        matched_keys = get_entry_data_colnames(patterns, entry)
               
        if set(matched_keys.keys()) == set(must_get) or 'refractive_index' in must_get:
            entry_data = entry.data
            num_data_points = entry_data.shape[0]
            smile = entry.components[0].smiles
            ion_smiles = separate_cat_from_an_similes(smile)
            if len(ion_smiles) == 2:
                cation_smile, anion_smile = ion_smiles
            else:
                cation_smile, anion_smile = None , None
            data['id'].extend([entry.id] * num_data_points)
            data['references'].extend([format_reference_robust(entry.ref.full)] * num_data_points)
            data['il_name'].extend([entry.components[0].name]*num_data_points)
            data['il_smile'].extend([smile] * num_data_points)
            data['cation_smile'].extend([cation_smile] * num_data_points)
            data['anion_smile'].extend([anion_smile] * num_data_points)
            data['molecular_weight_g'].extend([entry.components[0].mw] * num_data_points)
            for key, val in matched_keys.items():
                    data[key].extend(entry_data[matched_keys[key]].values.tolist())
            if 'refractive_index' in must_get and 'wavelength' not in list(matched_keys.keys()):
                data['wavelength'].extend([589]*num_data_points)
            if 'refractive_index' in must_get and 'pressure' not in list(matched_keys.keys()):
                data['pressure'].extend([101.325]*num_data_points)
                
                            
    df = pd.DataFrame(data)
    return df
    
def get_range(df, key):
    q25 = df[key].quantile(0.25)
    q75 = df[key].quantile(0.75)
    IQR = q75 - q25 
    low = q25 - 1.5*IQR
    high = q75 + 1.5*IQR 
    return (low, high)

def save_data_to_excel(columns, must_get, patterns, entries_id_cmp1, save=True):
    def get_first_existing(must_get, props):
        for p in props:
            if p in must_get:
                return must_get.index(p)
        return None  # or raise error
    
    df = get_database_data(columns, must_get, patterns, entries_id_cmp1)
    print(df.shape)
    p_id = must_get.index('pressure')
    t_id = must_get.index('temperature')
    err_id = must_get.index('error')
    prop_id = get_first_existing(must_get, ["viscosity", "density", "conductivity", "heat_capacity","refractive_index"])
    
    df[f"{must_get[t_id]}_K"] = df[must_get[t_id]]
    df[f"{must_get[p_id]}_KPa"] = df[must_get[p_id]]
    df[f"{must_get[p_id]}_MPa"] = df[must_get[p_id]]/1000
    if must_get[prop_id] == 'viscosity':
        df[f"{must_get[prop_id]}_Pa*s"] = df[must_get[prop_id]]
        df[f"{must_get[prop_id]}_cP"] = df[must_get[prop_id]]*1000
        df[f"{must_get[err_id]}_Pa*s"] = df[must_get[err_id]]
        df[f"{must_get[err_id]}_cP"] = df[must_get[err_id]]*1000
    if must_get[prop_id] == 'density':
        df[f"{must_get[prop_id]}_Kg*m-3"] = df[must_get[prop_id]]
        df[f"{must_get[prop_id]}_g-mL"] = df[must_get[prop_id]]/1000
        df[f"{must_get[err_id]}_g-mL"] = df[must_get[err_id]]/1000
        df[f"{must_get[err_id]}_Kg*m-3"] = df[must_get[err_id]]
    if must_get[prop_id] == 'conductivity':
        df[f"{must_get[prop_id]}_S-m"] = df[must_get[prop_id]]
        df[f"{must_get[prop_id]}_mS-m"] = df[must_get[prop_id]]*1000
        df[f"{must_get[err_id]}_S-m"] = df[must_get[err_id]]
        df[f"{must_get[err_id]}_mS-m"] = df[must_get[err_id]]*1000
    if must_get[prop_id] == 'heat_capacity':
        df[f"{must_get[prop_id]}_J-K*mol"] = df[must_get[prop_id]]
        df[f"{must_get[err_id]}_J-K*mol"] = df[must_get[err_id]]
    if must_get[prop_id] == 'refractive_index':
        df['wavelength_nm'] = df['wavelength']
    
    df_needed = df[(df[must_get[prop_id]].between(*get_range(df, must_get[prop_id]))) &  
                             (df[must_get[t_id]].between(*get_range(df, must_get[t_id]))) & 
                            (df[must_get[p_id]].between(80, 500))].copy()  
    
    if must_get[prop_id] == 'refractive_index':
        df = df.drop(columns=[must_get[p_id], must_get[t_id]])
        df_needed = df_needed.drop(columns=[must_get[p_id], must_get[t_id]])
    else:
        df = df.drop(columns=[must_get[prop_id], must_get[p_id], must_get[t_id], must_get[err_id]])
        df_needed = df_needed.drop(columns=[must_get[prop_id], must_get[p_id], must_get[t_id], must_get[err_id]])

    if 'wavelength' in list(df.columns):
        df = df.drop(columns=['wavelength'])
        df_needed = df_needed.drop(columns=['wavelength'])
    
    print(f"Data statistics from database\n{df.describe()}\n")
    print(f"Data statistics from database after filtering and transformations\n{df_needed.describe()}")
    
    def save_df_to_excel(df, filename=os.path.join(DATA, f"{must_get[prop_id]}-ambient.xlsx"), sheet_name=f"{must_get[prop_id]}", index=False):
        df.to_excel(filename, sheet_name=sheet_name, index=index)
        # 2. Open workbook with openpyxl for formatting
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        # 3. Format header row (bold)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
        # 4. Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_length = len(str(cell.value))
                    if val_length > max_length:
                        max_length = val_length
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # padding
    
        # 5. Save formatted Excel
        wb.save(filename)
        print(f"DataFrame saved to '{filename}' with formatting applied.")
    if save:
        save_df_to_excel(df_needed, filename=os.path.join(DATA, f"{must_get[prop_id]}_ambient.xlsx"), sheet_name=f"{must_get[prop_id]}", index=False)
        save_df_to_excel(df, filename=os.path.join(DATA,f"{must_get[prop_id]}_all_condition_range.xlsx"), sheet_name=f"{must_get[prop_id]}", index=False)
    return df_needed, df


'''Extracting and saving Viscosity Data'''
patterns = {
    "temperature": re.compile(r"^temperature", re.IGNORECASE),
    "viscosity": re.compile(r"^(viscosity)\b", re.IGNORECASE), #re.compile(r"^(viscosity|kinematic viscosity)\b", re.IGNORECASE)
    "pressure": re.compile(r"pressure", re.IGNORECASE),
    "error": re.compile(r"^error", re.IGNORECASE),
}
columns = ['id', 'references', 'il_name', 'il_smile', 'cation_smile', 'anion_smile', 'temperature', 'pressure','viscosity','error','molecular_weight_g']
must_get = ['temperature', 'pressure', 'viscosity', 'error']

df_visc, df_visc_all = save_data_to_excel(columns, must_get, patterns, entries_id_cmp1_visc)


'''Extracting and saving Density Data'''
patterns = {
    "temperature": re.compile(r"^temperature", re.IGNORECASE),
    "density": re.compile(r"^(density|specific density)\b", re.IGNORECASE), 
    "pressure": re.compile(r"pressure", re.IGNORECASE),
    "error": re.compile(r"^error", re.IGNORECASE),
}
columns = ['id', 'references', 'il_name', 'il_smile', 'cation_smile', 'anion_smile', 'temperature', 'pressure','density','error','molecular_weight_g']
must_get = ['temperature', 'pressure', 'density', 'error']

df_dens, df_dens_all = save_data_to_excel(columns, must_get, patterns, entries_id_cmp1_dens)

'''Extracting and saving Conductivity Data'''
patterns = {
    "temperature": re.compile(r"^temperature", re.IGNORECASE),
    "conductivity": re.compile(r"^(Electrical conductivity)\b", re.IGNORECASE), 
    "pressure": re.compile(r"^pressure", re.IGNORECASE),
    "error": re.compile(r"^error", re.IGNORECASE),
}
columns = ['id', 'references', 'il_name', 'il_smile', 'cation_smile', 'anion_smile', 'temperature','pressure', 'conductivity','error','molecular_weight_g']
must_get = ['temperature', 'pressure', 'conductivity', 'error']

df_cond, df_cond_all = save_data_to_excel(columns, must_get, patterns, entries_id_cmp1_cond)


'''Extracting and saving Refractive index Data'''
patterns = {
    "temperature": re.compile(r"^temperature", re.IGNORECASE),
    "refractive_index": re.compile(r"^(Refractive index)\b", re.IGNORECASE), 
    "pressure": re.compile(r"^pressure", re.IGNORECASE),
    "error": re.compile(r"^error", re.IGNORECASE),
    "wavelength": re.compile(r"^wavelength", re.IGNORECASE),
}
columns = ['id', 'references', 'il_name', 'il_smile', 'cation_smile', 'anion_smile', 'temperature','pressure','wavelength', 'refractive_index','error','molecular_weight_g']
must_get = ['temperature', 'pressure', 'wavelength', 'refractive_index', 'error']

df_refrac, df_refrac_all = save_data_to_excel(columns, must_get, patterns, entries_id_cmp1_refrac)

'''
You can also wish to visualize ILs componenets
'''


'''Here are for the preprocesing of the Deep Eutectic Solvents Dataset
Source: https://www.sciencedirect.com/science/article/pii/S0378381222002916?via%3Dihub or https://pubs.acs.org/doi/10.1021/acs.iecr.4c03610
'''

# A function to save data to excel and adjust column width

def save_df_to_excel(df, filename=None, sheet_name=f"Data", index=False):
    df.to_excel(filename, sheet_name=sheet_name, index=index)
    # 2. Open workbook with openpyxl for formatting
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    # 3. Format header row (bold)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 4. Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_length = len(str(cell.value))
                if val_length > max_length:
                    max_length = val_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # padding

    # 5. Save formatted Excel
    wb.save(filename)


'''
Here I extract and save the DES data accordingly
 '''
des_file = os.path.join(DATA, 'des-raw-density.xlsx')
df_des = pd.read_excel(des_file)
unique_HBA = df_des['HBA'].unique().tolist()
unique_HBD = df_des['HBD'].unique().tolist()

''' A dictionary of unique HBA and HBD as keys with smile as values '''
des_smiles_dict = {
    "Acetyl choline chloride": "CC(=O)OCC[N+](C)(C)C.[Cl-]",
    "Allyl triphenylphosphonium bromide": "C=CC[P+](C1=CC=CC=C1)(C2=CC=CC=C2)C3=CC=CC=C3.[Br-]",
    "Benzyl tripropyl ammonium Chloride": "CCC[N+](CCC)(CCC)CC1=CC=CC=C1.[Cl-]",
    "Betaine": "C[N+](C)(C)CC(=O)[O-]",
    "Benzyldimethyl(2-hydroxyethyl) ammonium chloride": "C[N+](C)(CCO)Cc1ccccc1.[Cl-]",
    "Choline chloride": "C[N+](C)(C)CCO.[Cl-]",
    "Citric acid": "C(C(=O)O)C(CC(=O)O)(C(=O)O)O",
    "D-Glucose": "C([C@@H]1[C@H]([C@@H]([C@H](C(O1)O)O)O)O)O",
    "Diethylamine hydrochloride": "CCN(CC)CC.Cl",
    "L-proline": "C1CC(NC1)C(=O)O",
    "Methyl triphenylphosphonium bromide": "C[P+](C1=CC=CC=C1)(C2=CC=CC=C2)C3=CC=CC=C3.[Br-]",
    "N,N diethylenethanol ammonium chloride": "C[N+](C)(CCO)CCO.[Cl-]",
    "Tetrabutylammonium chloride": "CCCC[N+](CCCC)(CCCC)(CCCC).[Cl-]",
    "Tetraethylammonium chloride": "CC[N+](CC)(CC)CC.[Cl-]",
    "Tetraethylammonium bromide": "CC[N+](CC)(CC)CC.[Br-]",
    "Tetrahexylammonium bromide": "CCCCCC[N+](CCCCCC)(CCCCCC)(CCCCCC).[Br-]",
    "Trimethylglicine": "C[N+](C)(C)CC(=O)[O-]",
    "1,2,4-triazole": "N1C=NC=N1",
    "D-Fructose": "C(C(C(C(C(CO)O)O)O)O)O",
    "D-Mannose": "C(C(C(C(C(CO)O)O)O)O)O",
    "D-Ribose": "C(C(C(C(CO)O)O)O)O",
    "D-Xylose": "C(C(C(C(CO)O)O)O)O",
    "Guaiacol": "COC1=CC=C(C=C1)O",
    "Imidazole": "c1cnc[nH]1",
    "Levulinic acid": "CC(=O)CCC(=O)O",
    "Diethylene glycol": "OCCOCCO",
    "Triethylene glycol": "OCCOCCOCCO",
    "Ethylene Glycol": "OCCO",
    "Glycerol": "C(C(CO)O)O",
    "Lactic acid": "CC(O)C(=O)O",
    "Oxalic acid": "C(=O)C(=O)O",
    "Phenol": "c1ccccc1O",
    "1,2-propanediol": "CC(CO)O",
    "1,4-butanediol": "OCCCCO",
    "2,3-butanediol": "CC(O)C(C)O",
    "Acetamide": "CC(=O)N",
    "D-Sorbitol": "C(C(C(C(C(CO)O)O)O)O)O",
    "D-Sucrose": "C1[C@@H]([C@H]([C@@H]([C@H](O1)OC2C(C(C(C(O2)CO)O)O)O)O)O)O",
    "Glutaric acid": "O=C(CCCCC(=O)O)O",
    "Glycolic acid": "C(C(=O)O)O",
    "Malonic acid": "O=C(C(=O)O)C(=O)O",
    "Nfurfuryl alcohol": "OCC1=CC=CO1",
    "O-Cresol": "Cc1ccc(O)cc1",
    "P-Chlorophenol": "c1cc(Cl)ccc1O",
    "P-Cresol": "Cc1ccc(O)cc1",
    "p-Toluenesulfonic acid": "Cc1ccc(cc1)S(=O)(=O)O",
    "Tartaric acid": "C(C(C(=O)O)O)C(=O)O",
    "Urea": "C(N)N=O",
    "Xylitol": "C(C(C(C(CO)O)O)O)CO",
    "Asinine": "C(C(=O)O)N",
    "Aspartic acid": "C(C(C(=O)O)N)C(=O)O",
    "Glutamic acid": "C(CC(=O)O)C(C(=O)O)N",
    "Phenylacetic acid": "c1ccccc1CC(=O)O",
    "Propionic acid": "CC(C(=O)O)",
    "2-Chloro benzoic acid": "ClC1=CC=CC=C1C(=O)O",
    "Benzoic acid": "c1ccccc1C(=O)O",
    "Mandelic acid": "c1ccccc1C(O)C(=O)O"
}


''' Making des_smile, hba_smile, and hbd_smile columns'''

des_dict = {'des_smile':[], 'hba_smile':[], 'hbd_smile':[]}

des_smiles_dict = {k.lower(): v for k,v in des_smiles_dict.items()}

for hba, hbd in zip(df_des['HBA'].values, df_des['HBD'].values):
    hba_smi = des_smiles_dict[hba.lower()]
    hbd_smi = des_smiles_dict[hbd.lower()]
    des_smi = f"{hba_smi}.{hbd_smi}"
    des_dict['des_smile'].append(des_smi)
    des_dict['hba_smile'].append(hba_smi)
    des_dict['hbd_smile'].append(hbd_smi)

df_des = pd.concat([df_des, pd.DataFrame(des_dict)], axis=1, ignore_index=False)


''' Now let us make other features including assumption of pressure to be 101.325 KPa since it was not providede in the dataset'''

df_des[['n_HBA', 'n_HBD']] = (df_des['HBA:HBD'].str.replace(',', ';')  # unify separators
    .str.split(';', expand=True).astype(float))
df_des['HBA_HBD_ratio'] = df_des['n_HBA'] / df_des['n_HBD']
df_des['pressure_KPa'] = 101.325
df_des['pressure_MPa'] = df_des['pressure_KPa']/1000
df_des[['pressure_KPa', 'pressure_MPa']] = df_des[['pressure_KPa', 'pressure_MPa']].astype('float')

columns = df_des.columns.values.tolist()
print(columns)

rename_columns = {'T(K)':'temperature_K', 'Density (kg/m3)':'density_Kg*m-3'}
df_des = df_des.rename(columns=rename_columns)
df_des = df_des.drop(columns=['AARD%'])
df_des['mol'] = df_des['des_smile'].apply(lambda x: Chem.MolFromSmiles(x) if x else None)
save_df_to_excel(df_des, filename=os.path.join(DATA, 'des-density_ambient.xlsx'), sheet_name=f"Data", index=False)